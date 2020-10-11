# pip install openpyxl
import openpyxl
import os
dir = os.getcwd()
print(dir)
os.chdir('C:/Users/Vivek Kumar/Desktop')
dir = os.getcwd()
print(dir)
# print(os.listdir())

workbook=openpyxl.load_workbook('sample.xlsx')
# print(workbook.sheet())
# print(type(workbook))
# print(workbook.get_sheet_names())
# print(workbook.A1())

sheet1=workbook.get_sheet_by_name('Sheet1')
cell=sheet1['D4']
print(cell.value)

print(sheet1.cell(row=1,column=2))

for i in range(1,8):
    print(i,sheet1.cell(row=2,column=i).value)