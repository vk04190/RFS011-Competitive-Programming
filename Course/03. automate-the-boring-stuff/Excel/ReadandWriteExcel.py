import os
import openpyxl
wb = openpyxl.Workbook()

# print(wb)
# print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet')
print(sheet['A1'].value)
sheet['A1'] = 'abc'
print(sheet['A1'].value)

sheet2=wb.create_sheet()

sheet2.title="Vivek"

wb.create_sheet(index=6,title="Betwwen")
for i in range(10):
    wb.create_sheet(index=i,title="Test Sheet Number %s" % (i))

os.chdir("c:/Users/Vivek Kumar/Desktop/")
wb.save('Example.xlsx')
print(wb.get_sheet_names())